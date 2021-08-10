
#%%
#%matplotlib inline
#from common.PlanetarySystem import PlanetarySystem
import PSEUGA.src.Thesis as helpers
import PSEUGA.common.Constants as const
import PSEUGA.common.PlanetarySystem
import PSEUGA.src.inputhandling as input


import lightkurve as lk
import numpy as np
import random

import sys, getopt
import os
import time

import matplotlib

import csv
import openpyxl as xl
from openpyxl.formatting.rule import ColorScaleRule
import jsonpickle




def getInputs():
    global inputFile
    global curveOutputFile
    global dataOutputFile
    global populationOutputFile
    global populationOutputExcel

    #try:
    #    opts, args = getopt.getopt(sys.argv[1:], "i:c:d:")
    #except getopt.GetoptError:
    #    print('viewer.py -i <inputfile> -c <lightcurveoutputfile> -d <dataoutputfile>')
    #for opt, arg in opts:
    #    if opt == '-i':
    #        inputFile = arg
    #    elif opt == '-c':
    #        curveOutputFile = arg
    #    elif opt == '-d':
    #        dataOutputFile = arg 
    
    prefix = "testing"

    inputFile = "DefaultFileTIC307210830C.fits"
    curveOutputFile = prefix + "curve.fits"
    dataOutputFile = prefix + "data.txt"
    populationOutputFile = prefix + "pop.csv"
    populationOutputExcel = prefix + ".xlsx"

    print('-------------------------------------------------------')
    print("Input: %s" %inputFile)
    print("Curve: %s" %curveOutputFile)
    print("Data : %s" %dataOutputFile)
    print("pop : %s" %populationOutputFile)
    const.skippedTimesteps = 1


def openFiles():
    helpers.targetCurve = lk.read(inputFile)
    #result = lk.read(curveOutputFile)
    with open(dataOutputFile) as f:
        individualStrings = f.read().splitlines()
    individual = [i for i in range(5 + const.ATTRPERPLANET * const.MAXPLANETS)]
    for i in range(const.MAXPLANETS):
        individual[i * const.ATTRPERPLANET + const.RADIUS] = float(individualStrings[i * const.ATTRPERPLANET + const.RADIUS])
        individual[i * const.ATTRPERPLANET + const.ECC] = float(individualStrings[i * const.ATTRPERPLANET + const.ECC])
        individual[i * const.ATTRPERPLANET + const.SMA] = float(individualStrings[i * const.ATTRPERPLANET + const.SMA])
        individual[i * const.ATTRPERPLANET + const.INC] = float(individualStrings[i * const.ATTRPERPLANET + const.INC])
        individual[i * const.ATTRPERPLANET + const.LOAN] = float(individualStrings[i * const.ATTRPERPLANET + const.LOAN])
        individual[i * const.ATTRPERPLANET + const.AOP] = float(individualStrings[i * const.ATTRPERPLANET + const.AOP])
        individual[i * const.ATTRPERPLANET + const.MA] = float(individualStrings[i * const.ATTRPERPLANET + const.MA])
        
    individual[const.NUMPLANETS] = int(individualStrings[const.NUMPLANETS])
    individual[const.STARRADIUS] = float(individualStrings[const.STARRADIUS])
    individual[const.STARMASS] = float(individualStrings[const.STARMASS])
    individual[const.STARBASEFLUX] = float(individualStrings[const.STARBASEFLUX])
    individual[const.DISTANCE] = float(individualStrings[const.DISTANCE])


    print('-------------------------------------------------------')
    inputVars = const.DISTANCE + 3
    print("Number of threads:................... %s" %individualStrings[inputVars])
    print("Skipped timesteps:................... %s" %individualStrings[inputVars + 1])
    print("Number of generations:............... %s" %individualStrings[inputVars + 2])
    print("Number of individuals:............... %s" %individualStrings[inputVars + 3])

    print('-------------------------------------------------------')
    recomboVars = inputVars + 6
    print("Crossover prob:...................... %s" %individualStrings[recomboVars])
    print("Mutation prob:....................... %s" %individualStrings[recomboVars + 1])
    
    print('-------------------------------------------------------')
    statsVars = recomboVars + 4
    print("Mean fitness:........................ %s" %individualStrings[statsVars])
    print("Fitness std:......................... %s" %individualStrings[statsVars+1])
    print("Min fitness:......................... %s" %individualStrings[statsVars+2])
    print("Max fitness:......................... %s" %individualStrings[statsVars+3])

    wb = xl.Workbook()
    ws = wb.active
    for planet in range(const.MAXPLANETS):
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.RADIUS + 1)
        cell.value = 'Radius (km)'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.ECC + 1)
        cell.value = 'Eccentricity'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.SMA + 1)
        cell.value = 'Semi-major Axis (km)'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.INC + 1)
        cell.value = 'Inclination (rads)'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.LOAN + 1)
        cell.value = 'Long. of Asc. Node (rads)'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.AOP + 1)
        cell.value = 'Arg. of Periapse (rads)'
        cell = ws.cell(row=1,column=planet * const.ATTRPERPLANET + const.MA + 1)
        cell.value = 'Mean Anomaly (rads)'
    
    
    cell = ws.cell(row=1,column=const.NUMPLANETS + 1)
    cell.value = 'Num. of Planets'
    cell = ws.cell(row=1,column=const.STARRADIUS + 1)
    cell.value = 'Star Radius (km)'
    cell = ws.cell(row=1,column=const.STARMASS + 1)
    cell.value = 'Star Mass (kg)'
    cell = ws.cell(row=1,column=const.STARBASEFLUX + 1)
    cell.value = 'Star Base Flux (e/s)'
    cell = ws.cell(row=1,column=const.DISTANCE + 1)
    cell.value = 'Star Distance (km)'

    numRows = 1
    rowlen = 0
    with open(populationOutputFile) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            rowNumbers = list(map(float,row))
            rowlen = len(row)
            ws.append(rowNumbers)
            numRows = numRows + 1

    for i in range(1,rowlen+1):
        cell = ws.cell(row=numRows+2,column=i)
        cell.value = '=AVERAGE({0}2:{0}{1})'.format(xl.utils.cell.get_column_letter(i), numRows)
        cell = ws.cell(row=numRows+3,column=i)
        cell.value = '=STDEV({0}2:{0}{1})'.format(xl.utils.cell.get_column_letter(i), numRows)
        cell = ws.cell(row=numRows+4,column=i)
        cell.value = '={0}{1} / {0}{2}'.format(xl.utils.cell.get_column_letter(i), numRows+3, numRows+2)

    ws.conditional_formatting.add('A{0}:{1}{0}'.format(numRows+4,xl.utils.cell.get_column_letter(rowlen+1)), ColorScaleRule(start_type='min', start_color='FFAAAA', end_type='max', end_color='AAFFAA'))

    wb.save(populationOutputExcel)
    print("saved")

    result = helpers.generateLightcurve(individual)
    result.plot()
    helpers.targetCurve.plot()
    flattened = helpers.targetCurve.flatten()
    flattened.plot()
    plotAvg = result[1]
    print(plotAvg)
    
def testPickle():
    indiv = ''
    with open('test.json') as f:
        reader = f.read()
        indiv = jsonpickle.decode(reader)
    print(f"Found {indiv.numActivePlanets} active planets")

def testOpenLightcurve(filepath):
    kurve = lk.read(filepath)
    return kurve

def main():
    #getInputs()
    #openFiles()
    paths = input.getIOFromInput(sys.argv, False)
    print(paths)

if __name__ == "__main__":
    main()
# %%
